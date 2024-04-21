from pathlib import Path
from tkinter import filedialog
import pandas as pd
import numpy as np
import openpyxl
import os
import time
from openpyxl import load_workbook

# Funcion para comprobar la primera fila del archivo de texto
def verificar_primera_fila(archivo, cadena):
    with open(archivo, 'r') as file:
        primera_fila = file.readline().strip()
        return primera_fila.startswith(cadena)
    

# Mostrar el diálogo para abrir un archivo.
filename = filedialog.askopenfilename(initialdir='S:/LDDAT/PRUEBAS LIQUIDACION')

initialdir='S:/LDDAT/PRUEBAS LIQUIDACION'
# Abro el archivo en modo lectura
with open(filename, 'r') as archivo:
    # Leo y guardo el contenido actual del archivo
    contenido_original = archivo.read()
ruta = os.path.split(filename)   
nombre_archivo_txt = ruta[1]
nombre_excel = nombre_archivo_txt[4:11]

# Creo una nueva fila con el encabezado de cada columna
fila_encabezado = "CUIT;ORGANISMO;AÑO;MES;ANEXO;AGRUP;CATEGORIA;CLASE;ITEM;COD-13-MAE;COD-23-MAE;COD-27-MAE;PLANTA;LOCALIDAD;DESTINO;CUIL;APELLIDO Y NOMBRE;LEGAJO;CODIGO;DESCRIPCION;IMPORTE\n"
contenido_modificado = fila_encabezado + contenido_original

# Verifico si la primera fila comienza con CUIT no agrega la fila_encabezado
cadena_especifica = 'CUIT'  
if verificar_primera_fila(filename, cadena_especifica):
    print("La primera fila comienza con la cadena especificada.")
else:
    # Cuando no exista la nueva_fila en la primera linea del archivo txt, inserta la nueva_fila que contiene los encabezados al archivo txt
    with open(filename, 'w') as archivo_escritura:
        # Escribir el contenido modificado en el archivo
        archivo_escritura.write(contenido_modificado)

# Convierto el archivo txt en dataframe separado por ; y , como separador decimal
vertical = pd.read_csv(filename, sep = ";", decimal=",", encoding= "unicode_escape", )


# Definir la posición específica donde se agregará la nueva columna
posicion_cuit_organismo = 2
posicion_codigo_detalle = 21


# Realizar la concatenación de las columnas
vertical['CUIT-ORGANISMO'] = vertical['CUIT'].astype(str) +'-'+ vertical['ORGANISMO'].astype(str)
vertical['CODIGO-DESCRIPCION'] = vertical['CODIGO'].astype(str) +'-'+ vertical['DESCRIPCION'].astype(str)


# Mover la nueva columna a la posición deseada
columnas = list(vertical.columns)
columnas = columnas[:posicion_cuit_organismo] + ['CUIT-ORGANISMO'] + columnas[posicion_cuit_organismo:-1]
columnas = columnas[:posicion_codigo_detalle] + ['CODIGO-DESCRIPCION'] + columnas[posicion_codigo_detalle:-1]
vertical= vertical[columnas]
#vertical= vertical[columnas2]

# Convierto la columna importe a float64
vertical['IMPORTE'] = vertical['IMPORTE'].astype('float64')


# Creo el archivo excel de salida
archivo_excel = vertical.to_excel(f'S:/LDDAT/PRUEBAS LIQUIDACION/totales_por_organismos_cobol-{nombre_excel}.xlsx',index=False)

# ruta_archivo = os.path.abspath(archivo_excel.name)
# path = f'{initialdir}/totales_por_organismos_cobol-{nombre_excel}.XLSX'

# # Abro el archivo Excel ya creado
# libro_trabajo = openpyxl.load_workbook(path)


# # Agrego una segunda hoja al excel
# segunda_hoja = libro_trabajo.create_sheet(title='Tabla Dinámica')

# # Creo una tabla dinámica en la segunda hoja
# tabla_dinamica = vertical.pivot_table(index="CODIGO-DESCRIPCION", columns="CUIT-ORGANISMO",values="IMPORTE",aggfunc="sum") #  para mostrar los datos
# tabla_dinamica.to_excel(f'{nombre_excel}.XLSX', sheet_name='Tabla Dinámica')

# # Guardo los cambios
# libro_trabajo.save(archivo_excel)



# Muestro proceso finalizado correctamente
print(f"Archivo de Excel creado correctamente en S:/LDDAT/PRUEBAS LIQUIDACION/tabladinamica-{nombre_excel}.xlsx'")