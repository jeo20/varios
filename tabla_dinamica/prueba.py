import pandas as pd
from openpyxl import load_workbook

# Leer el archivo Excel
datos = pd.read_excel("D:\\python\\varios-jeo\\tabla_dinamica\\liq-9924-26.xlsx", sheet_name="Sheet1")

# Crear una tabla dinámica
wb = load_workbook("D:\\python\\varios-jeo\\tabla_dinamica\\liq-9924-26.xlsx")
ws = wb["Hoja2"]  # Crear la tabla dinámica en la hoja "Hoja2"
pt = pd.pivot_table(ws)
pt.add_data(ws["A1:C" + str(datos.shape[0])])
pt.add_row_field("Columna1")
pt.add_column_field("Columna2")
pt.add_data_field("Columna3", sum)
pt.data_only = True
ws.append(pt)

# Guardar el archivo Excel con la tabla dinámica
wb.save("datos_con_tabla_dinamica.xlsx")
